import argparse
from defusedxml import ElementTree as ET
import pandas as pd
import os
import logging
from openpyxl import load_workbook
from pymongo import MongoClient
from pymongo.errors import ServerSelectionTimeoutError
from pathlib import Path

# Configure logging
from logging_config import setup_logging
logger = setup_logging()



def sanitize_for_log(value):
    """Sanitize input for logging to prevent log injection attacks"""
    if value is None:
        return "None"
    # Convert to string and remove/replace dangerous characters
    sanitized = str(value).replace('\n', '\\n').replace('\r', '\\r').replace('\t', '\\t')
    # Remove other control characters that could be used for injection
    sanitized = ''.join(char for char in sanitized if ord(char) >= 32 or char in '\t\n\r')
    # Limit length to prevent log flooding
    return sanitized[:200] + '...' if len(sanitized) > 200 else sanitized

def _validate_safe_path(file_path, base_path=None):
    """Validate file path to prevent path traversal attacks"""
    try:
        # Convert to Path object and resolve
        path = Path(file_path).resolve()
        
        # If base_path provided, ensure file is within base directory
        if base_path:
            base = Path(base_path).resolve()
            try:
                path.relative_to(base)
            except ValueError:
                return False
        
        # Check for path traversal after resolution
        if '..' in str(path) or not path.is_absolute():
            return False
                
        return True
    except (OSError, ValueError):
        return False


def calc_pace(total_time_s, distance_m):
    try:
        time_val = float(total_time_s) if total_time_s is not None else 0
        dist_val = float(distance_m) if distance_m is not None else 0
        if time_val > 0 and dist_val > 0:
            return (time_val / (dist_val / 1000.0)) / 60.0
    except (ValueError, TypeError):
        logger.debug(f"Invalid pace calculation inputs: time={sanitize_for_log(total_time_s)}, distance={sanitize_for_log(distance_m)}")

    return None


def _extract_float_from_element(element):
    """Extract float value from XML element with error handling"""
    if element is not None and element.text:
        try:
            return float(element.text)
        except ValueError:
            pass
    return None


def _extract_lap_data(lap, ns):
    """Extract lap-level data from XML lap element"""
    # Validate namespace to prevent injection
    if not isinstance(ns, dict) or "tcx" not in ns:
        raise ValueError("Invalid namespace provided")
    
    lap_start = lap.attrib.get("StartTime")
    lap_total_time = lap.find("tcx:TotalTimeSeconds", ns)
    lap_distance = lap.find("tcx:DistanceMeters", ns)
    
    total_time_s = _extract_float_from_element(lap_total_time)
    distance_m = _extract_float_from_element(lap_distance)
    pace = calc_pace(total_time_s, distance_m)
    
    return lap_start, total_time_s, distance_m, pace


def _extract_trackpoint_data(tp, ns):
    """Extract trackpoint-level data from XML trackpoint element"""
    # Validate namespace to prevent injection
    if not isinstance(ns, dict) or "tcx" not in ns:
        raise ValueError("Invalid namespace provided")
        
    data = {
        "Time": None,
        "Latitude": None,
        "Longitude": None,
        "Altitude_m": None,
        "Distance_m": None,
    }
    
    # Extract time
    time_elem = tp.find("tcx:Time", ns)
    if time_elem is not None:
        data["Time"] = time_elem.text
    
    # Extract position (latitude/longitude)
    pos_elem = tp.find("tcx:Position", ns)
    if pos_elem is not None:
        lat_elem = pos_elem.find("tcx:LatitudeDegrees", ns)
        lon_elem = pos_elem.find("tcx:LongitudeDegrees", ns)
        data["Latitude"] = _extract_float_from_element(lat_elem)
        data["Longitude"] = _extract_float_from_element(lon_elem)
    
    # Extract altitude and distance
    data["Altitude_m"] = _extract_float_from_element(tp.find("tcx:AltitudeMeters", ns))
    data["Distance_m"] = _extract_float_from_element(tp.find("tcx:DistanceMeters", ns))
    
    return data


def parse_tcx_detailed(tcx_file):
    try:
        ns = {"tcx": "http://www.garmin.com/xmlschemas/TrainingCenterDatabase/v2"}
        # Validate namespace
        if not isinstance(ns, dict) or "tcx" not in ns:
            raise ValueError("Invalid namespace configuration")
        tree = ET.parse(tcx_file)
        root = tree.getroot()
        logger.info(f"Successfully parsed TCX file for detailed analysis: {sanitize_for_log(tcx_file)}")
    except ET.ParseError as e:
        logger.error(f"XML parsing error in {sanitize_for_log(tcx_file)}: {sanitize_for_log(e)}")
        raise ValueError(f"Invalid XML file: {e}")
    except FileNotFoundError as e:
        logger.error(f"TCX file not found: {sanitize_for_log(tcx_file)}")
        raise
    except Exception as e:
        logger.error(f"Unexpected error parsing {sanitize_for_log(tcx_file)}: {sanitize_for_log(e)}")
        raise

    rows = []
    lap_counter = 0

    for lap in root.findall(".//tcx:Lap", ns):
        lap_counter += 1
        lap_start, total_time_s, distance_m, pace = _extract_lap_data(lap, ns)

        for tp in lap.findall(".//tcx:Trackpoint", ns):
            trackpoint_data = _extract_trackpoint_data(tp, ns)
            
            row = {
                "LapNumber": lap_counter,
                "LapStartTime": lap_start,
                "LapTotalTime_s": total_time_s,
                "LapDistance_m": distance_m,
                "Pace_min_per_km": pace,
                **trackpoint_data
            }
            rows.append(row)

    return pd.DataFrame(rows)


def _parse_tcx_file(tcx_file, operation="analysis"):
    """Common XML parsing logic for TCX files"""
    try:
        ns = {"tcx": "http://www.garmin.com/xmlschemas/TrainingCenterDatabase/v2"}
        # Validate namespace
        if not isinstance(ns, dict) or "tcx" not in ns:
            raise ValueError("Invalid namespace configuration")
        tree = ET.parse(tcx_file)
        root = tree.getroot()
        logger.info(f"Successfully parsed TCX file for {operation}: {sanitize_for_log(tcx_file)}")
        return root, ns
    except ET.ParseError as e:
        logger.error(f"XML parsing error in {sanitize_for_log(tcx_file)}: {sanitize_for_log(e)}")
        raise ValueError(f"Invalid XML file: {e}")
    except FileNotFoundError as e:
        logger.error(f"TCX file not found: {sanitize_for_log(tcx_file)}")
        raise
    except Exception as e:
        logger.error(f"Unexpected error parsing {sanitize_for_log(tcx_file)}: {sanitize_for_log(e)}")
        raise

def parse_tcx_summary(tcx_file):
    root, ns = _parse_tcx_file(tcx_file, "summary analysis")

    rows = []
    lap_counter = 0

    for lap in root.findall(".//tcx:Lap", ns):
        lap_counter += 1
        lap_start, total_time_s, distance_m, pace = _extract_lap_data(lap, ns)

        rows.append({
            "LapNumber": lap_counter,
            "LapStartTime": lap_start,
            "LapTotalTime_s": total_time_s,
            "LapDistance_m": distance_m,
            "Pace_min_per_km": pace,
        })

    return pd.DataFrame(rows)


def get_first_lap_date(tcx_file):
    try:
        root, ns = _parse_tcx_file(tcx_file, "date extraction")
        first_lap = root.find(".//tcx:Lap", ns)
        if first_lap is not None and "StartTime" in first_lap.attrib:
            start_time = first_lap.attrib["StartTime"]
            # Safely split timestamp to extract date
            if "T" in start_time:
                date = start_time.split("T")[0]
                logger.debug(f"Extracted date {sanitize_for_log(date)} from {sanitize_for_log(tcx_file)}")
                return date
            else:
                logger.warning(f"Invalid timestamp format in {sanitize_for_log(tcx_file)}: {sanitize_for_log(start_time)}")
        else:
            logger.warning(f"No lap with StartTime found in {sanitize_for_log(tcx_file)}")
    except ET.ParseError as e:
        logger.error(f"XML parsing error while extracting date from {sanitize_for_log(tcx_file)}: {sanitize_for_log(e)}")
    except Exception as e:
        logger.error(f"Unexpected error extracting date from {sanitize_for_log(tcx_file)}: {sanitize_for_log(e)}")

    return "UnknownDate"


def write_to_excel(df, output_file, sheet_name):
    try:
        if not os.path.exists(output_file):
            with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
            logger.info(f"Created new Excel file: {sanitize_for_log(output_file)} with sheet: {sanitize_for_log(sheet_name)}")
        else:
            book = load_workbook(output_file)
            existing_sheets = book.sheetnames

            # Remove the sheet if it exists (case-insensitive)
            match_sheet = None
            for s in existing_sheets:
                if s.lower() == sheet_name.lower():
                    match_sheet = s
                    break

            if match_sheet:
                del book[match_sheet]
                book.save(output_file)  # Save immediately after deletion!
                logger.info(f"Replaced existing sheet: {sanitize_for_log(match_sheet)} in {sanitize_for_log(output_file)}")

            with pd.ExcelWriter(output_file, mode="a", engine="openpyxl") as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
            logger.info(f"Added sheet: {sanitize_for_log(sheet_name)} to existing file: {sanitize_for_log(output_file)}")
    except PermissionError as e:
        logger.error(f"Permission denied writing to Excel file {sanitize_for_log(output_file)}: {sanitize_for_log(e)}")
        raise
    except Exception as e:
        logger.error(f"Error writing to Excel file {sanitize_for_log(output_file)}: {sanitize_for_log(e)}")
        raise


def _sanitize_mongo_value(value):
    """Sanitize values for MongoDB queries to prevent NoSQL injection"""
    if value is None:
        return None
    # Only allow basic data types, reject complex objects that could contain operators
    if isinstance(value, (str, int, float, bool)):
        return value
    # Convert other types to string to prevent injection
    return str(value)

def push_to_mongo(df, collection, unique_keys):
    """
    Upsert each row in df to the mongo collection.
    unique_keys: list of column names for the unique query filter.
    """
    records = df.to_dict(orient="records")
    for rec in records:
        # Sanitize query values to prevent NoSQL injection
        query = {k: _sanitize_mongo_value(rec.get(k)) for k in unique_keys}
        # Sanitize record values
        sanitized_rec = {k: _sanitize_mongo_value(v) for k, v in rec.items()}
        collection.replace_one(query, sanitized_rec, upsert=True)


def process_file(tcx_file, args, mongo_client=None):
    logger.info(f"Starting processing of file: {sanitize_for_log(tcx_file)}")
    print(f"Processing {tcx_file}")

    date_str = get_first_lap_date(tcx_file)

    dfs_to_write = []
    dfs_to_mongo = []

    if args.mode in ("summary", "both"):
        df_summary = parse_tcx_summary(tcx_file)
        sheet_summary = f"{date_str}_summary"
        dfs_to_write.append((df_summary, sheet_summary))
        if mongo_client:
            dfs_to_mongo.append(("summary", df_summary))

    if args.mode in ("detailed", "both"):
        df_detail = parse_tcx_detailed(tcx_file)
        sheet_detail = f"{date_str}_detail"
        dfs_to_write.append((df_detail, sheet_detail))
        if mongo_client:
            dfs_to_mongo.append(("detailed", df_detail))

    # Write Excel sheets
    for df, sheet_name in dfs_to_write:
        write_to_excel(df, args.output, sheet_name)

    print(f"✅ Data written to Excel file '{args.output}'")

    # Push to MongoDB
    if mongo_client:
        db = mongo_client["RunningTracker"]
        for mode_name, df in dfs_to_mongo:
            collection = db[mode_name]

            # Determine unique keys for upsert based on mode
            if mode_name == "summary":
                unique_keys = ["LapStartTime", "LapNumber", "LapTotalTime_s", "LapDistance_m", "Pace_min_per_km"]
            else:  # detailed
                unique_keys = ["LapStartTime", "LapNumber", "Time"]

            # Add filename to each record for uniqueness and traceability
            filename = os.path.basename(tcx_file)
            if not _validate_safe_path(filename):
                filename = 'sanitized_file.tcx'
            df["_source_file"] = filename

            # Convert DataFrame to dicts with updated keys
            records = df.to_dict(orient="records")

            for rec in records:
                # Compose the filter query only with keys present in rec
                query = {k: _sanitize_mongo_value(rec.get(k)) for k in unique_keys if rec.get(k) is not None}
                # Include source file in query to avoid conflicts across files
                query["_source_file"] = _sanitize_mongo_value(rec["_source_file"])
                # Sanitize record values
                sanitized_rec = {k: _sanitize_mongo_value(v) for k, v in rec.items()}
                collection.replace_one(query, sanitized_rec, upsert=True)

        print("✅ Data pushed to MongoDB")


def _discover_tcx_files(input_path):
    """Discover TCX files from input path"""
    # Determine if input is file or folder
    if os.path.isfile(input_path):
        return [input_path]
    
    # Folder - get all .tcx files
    try:
        files = []
        for f in os.listdir(input_path):
            if f.lower().endswith(".tcx"):
                full_path = os.path.join(input_path, f)
                if _validate_safe_path(full_path, input_path) and os.path.isfile(full_path):
                    files.append(full_path)
                    
        if not files:
            print(f"No .tcx files found in folder '{input_path}'.")
            return []
        return files
    except (PermissionError, OSError) as e:
        logger.error(f"Directory access error: {sanitize_for_log(input_path)} - {sanitize_for_log(e)}")
        return []

def _setup_mongo_connection(args):
    """Setup MongoDB connection if requested"""
    if not args.mongo:
        return None
        
    try:
        mongo_client = MongoClient(args.mongo_uri, serverSelectionTimeoutMS=5000)
        # Trigger a server selection to verify connection
        mongo_client.server_info()
        logger.info(f"Successfully connected to MongoDB at {sanitize_for_log(args.mongo_uri)}")
        print("Connected to MongoDB")
        return mongo_client
    except ServerSelectionTimeoutError as e:
        logger.error(f"MongoDB connection timeout: {sanitize_for_log(args.mongo_uri)} - {sanitize_for_log(e)}")
        print("ERROR: Could not connect to MongoDB. Please check your connection.")
        return None
    except Exception as e:
        logger.error(f"Unexpected MongoDB connection error: {sanitize_for_log(e)}")
        print(f"ERROR: MongoDB connection failed: {e}")
        return None

def main():
    parser = argparse.ArgumentParser(
        description=(
            "Parse TCX file(s) (Garmin Training Center XML) into an Excel table and optionally MongoDB.\n"
            "Default export mode is 'both' unless --mode is specified."
        ),
        formatter_class=argparse.RawTextHelpFormatter,
    )

    parser.add_argument("input_path", help="Path to TCX file or folder containing TCX files.")
    parser.add_argument("--output", help="Path to Excel file. Default: tcx_data.xlsx", default="tcx_data.xlsx")
    parser.add_argument(
        "--mode",
        choices=["detailed", "summary", "both"],
        default="both",
        help="Export mode: summary, detailed, or both (default).",
    )
    parser.add_argument(
        "--mongo",
        action="store_true",
        help="Push the parsed data to MongoDB (requires 'pymongo'). MongoDB must be reachable at default localhost:27017.",
    )
    parser.add_argument(
        "--mongo-uri",
        default="mongodb://localhost:27017",
        help="MongoDB connection URI (default: mongodb://localhost:27017).",
    )

    args = parser.parse_args()

    # Validate input path
    if not os.path.exists(args.input_path):
        print(f"Input path '{args.input_path}' does not exist.")
        return

    mongo_client = _setup_mongo_connection(args)
    if args.mongo and mongo_client is None:
        return

    try:
        # Validate input path to prevent path traversal
        if not _validate_safe_path(args.input_path):
            print(f"Invalid or unsafe input path: '{args.input_path}'")
            return
            
        files = _discover_tcx_files(args.input_path)
        if not files:
            return

        for f in files:
            process_file(f, args, mongo_client)
    finally:
        if mongo_client:
            mongo_client.close()


if __name__ == "__main__":
    main()
